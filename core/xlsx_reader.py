# -*- coding: utf-8 -*-
"""
XLSX森林簿リーダー
openpyxlのread_onlyモードで大容量xlsxをストリーミング読込する。
"""
import logging
from typing import Iterator, Dict, Any, List

logger = logging.getLogger(__name__)


def read_xlsx(path: str, sheet_name: str = 'データ',
              progress_callback=None) -> Iterator[Dict[str, Any]]:
    """XLSXファイルをストリーミングで読み込み、行辞書を返す。

    Args:
        path: XLSXファイルパス
        sheet_name: シート名（デフォルト: 'データ'）
        progress_callback: 進捗コールバック (current: int, total: int) → None
                           load_workbook 完了直後に (0, total) で呼ばれ、
                           以降 5000 行ごと・最終行で (row_num, total) で呼ばれる。

    Yields:
        {カラム名: 値, ...} の辞書
    """
    import openpyxl

    # read_only=True は lxml.etree.iterparse（ジェネレータ）を使うため
    # wb.close() 後も _IterparseContext が GC 待ちで残る。
    # QThread 起動時に GIL 解放のタイミングで xmlDictFree が呼ばれ、
    # Windows の未初期化クリティカルセクションを取得しようとしてアクセス違反になる。
    # → read_only=False（デフォルト）で全量ロードし iterparse を避ける。
    wb = openpyxl.load_workbook(path, data_only=True)

    # シート名で検索、なければ最初のシート
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]
        logger.warning(f'シート "{sheet_name}" が見つかりません。"{wb.sheetnames[0]}" を使用します。')

    # load_workbook 完了 → 総行数が確定。コールバックで進捗バーの上限を通知する。
    total_rows = max((ws.max_row or 1) - 1, 0)  # ヘッダー行を除いたデータ行数
    if progress_callback:
        progress_callback(0, total_rows)

    headers = None
    row_num = 0
    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = [str(h).strip() if h is not None else f'col_{i}'
                       for i, h in enumerate(row)]
            continue

        row_dict = {}
        for i, val in enumerate(row):
            if i < len(headers):
                row_dict[headers[i]] = val

        row_num += 1
        if progress_callback and row_num % 5000 == 0:
            progress_callback(row_num, total_rows)

        yield row_dict

    if progress_callback:
        progress_callback(row_num, total_rows)  # 最終行の端数を確実に報告

    wb.close()



def get_cd_columns(headers: List[str]) -> List[str]:
    """ヘッダーリストからCD列名を抽出する。"""
    cd_cols = [h for h in headers if h.endswith('CD')]
    # 森林認証は CDで終わらないが変換対象
    if '森林認証' in headers:
        cd_cols.append('森林認証')
    # ゾーニング施業種
    if 'ゾーニング施業種' in headers:
        cd_cols.append('ゾーニング施業種')
    # 松林区分（CDで終わらないが変換対象）
    if '松林区分' in headers:
        cd_cols.append('松林区分')
    # ゾーニング機能（CDで終わらない5列）
    for h in headers:
        if h.startswith('ゾーニング機能_'):
            if h not in cd_cols:
                cd_cols.append(h)
    # 施業履歴関連
    for h in headers:
        if h.startswith('施業履歴_施業方法') or h.startswith('施業履歴_事業種類'):
            if h not in cd_cols:
                cd_cols.append(h)
    return cd_cols
